inputAlignment = "20170807_HIV1_ALL_2016_env_PRO.fasta" # An alignment file produced by AliView.
inputAlignmentPath = "/Users/newumuser/Desktop/silver4490/Desktop/Long V1 Envelope O-Glycosylation Story/V1 Insert Los Alamos Blast/Compendium 2016/netOGlyc-4.0.0.13/"

# Dependencies.
###############
# from openpyxl import Workbook
from operator import itemgetter
from openpyxl import Workbook
import numpy

# Get the file information.
###########################
input = open(inputAlignmentPath + inputAlignment, "r")
lines = input.readlines() # The readline() function returns a list of lines in the file.
input.close()

# Make a dictionary where the Fasta sequence name is the key, and the Fasta sequence is the value.
##################################################################################################
alignedSequences = {}
raw = []
i = 0
for line in lines:
	if ">" in line:
		raw.append(line)
	else:
		modified = line.strip('\n')
		raw.append(modified)
longString = ''.join(raw)

oldNameKey = {} ###This is to ensure that we use the correct name in the final output. This is because the O-glycosylation prediction removes certain characters.
lessStrings = longString.split('>')
del(lessStrings[0])
for item in lessStrings:
	curatedKeyValue = item.split('\n')
	newValue = curatedKeyValue[0].replace(".", "_")
	newestValue = newValue.replace("x", "")
	newlyestValue = newestValue.replace("X", "")
	finalValue = '>'+newlyestValue.upper()
	alignedSequences[finalValue] = curatedKeyValue[1]
	oldNameKey[finalValue] = '>'+curatedKeyValue[0]

##### Remove sequences with ambiguous characters
i = 0
j = 0
a = 0
b = 0
z = 0
newlyAlignedSequences = {}
removedSequences = []
for seq in alignedSequences:
	if "*" in alignedSequences[seq][:-1]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			z += 1
	elif "X" in alignedSequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			i += 1
	elif "#" in alignedSequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			j += 1
	elif "?" in alignedSequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			a += 1
	elif "$" in alignedSequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			b += 1
	else:
		newlyAlignedSequences[seq] = alignedSequences[seq]

print("There are", len(removedSequences), "sequences excluded from the Envelope Fasta because of unusual characters.")
print("There are", len(newlyAlignedSequences), "sequences in the modified Envelope Fasta.")

##########################################################
#########Sequences containing #, X, ?, $ Removed##########
##########################################################

### REFER TO newlyAlignedSequences Dictionary!

####################################################################################################
# Create a dictionary containing the sequence identifier as the key and the V1 Domain as the value #
# Calculate the length of each V1 domain ###########################################################
# This un-aligns the sequences by removing dashes and replacing them with nothing ##################
####################################################################################################
variableDomain = {}
for value in newlyAlignedSequences:
	fasta = newlyAlignedSequences[value]
	v1Domain = fasta[239:310]
	for letter in v1Domain:
		v1DomainCure = v1Domain.replace("-", "")
	variableDomain[value] = v1DomainCure

v1Length = {}
for value in variableDomain:
	length = len(variableDomain[value])
	v1Length[value] = length

serineThreonineCount = {}
for value in variableDomain:
	number=variableDomain[value].count('S')+variableDomain[value].count('T')
	serineThreonineCount[value] = number

curatedSequences = {}
for value in newlyAlignedSequences:
	removeDashes = newlyAlignedSequences[value].replace("-", "")
	curatedSequences[value] = removeDashes


###################################################################################
# Additionally, this writes the sequences to a fasta file for uplaod to NetOGlyc4 #
###################################################################################

# newFasta = open('UnalignedSequencesForNetOGlyc.fasta', 'w')
# for thing in curatedSequences:
# 	annoying = curatedSequences[thing]
# 	newFasta.write(thing + '\n')
# 	newFasta.write(annoying + '\n')

#################################################################
# Find the indices that border the V1 domain for each sequences #
# Write the indices to an empty text files                      #
# This will be used later to identify O-linked glycans in V1    #
#################################################################

relevantInfoDict = {}
v1DomainIndices = {}
for envelope in curatedSequences:
	beginning = curatedSequences[envelope].find(variableDomain[envelope])
	end = int(beginning) + int(len(variableDomain[envelope])-1)
	relevantInfoDict[envelope] = [variableDomain[envelope], len(variableDomain[envelope]), beginning, end]
	# print("The V1 domain of", envelope, "is", variableDomain[envelope], "and begins at AA", beginning, "and ends at AA", end)
	v1DomainIndices[envelope] = beginning+1, end+1

file = open('V1domain_BorderIndices.txt', 'w')
for line in v1DomainIndices:
	if v1Length[line] >= 0:
		shit = line + '\n' + str(v1DomainIndices[line][0]) + " " + str(v1DomainIndices[line][1]) + '\n'
		file.write(shit)
	else:
		pass

###############################################################################################
# Open O-glycosylation file, clean it, extract  relevant data   ###############################
# Modify O-glyc data to include only the sequence info from newlyAlignedSequences dict ########
###############################################################################################
inputAlignment = "20170831_MODIFIEDhiren_NetOGlycData.txt" # The O-glyc prediction software output.
inputAlignmentPath = "/Users/newumuser/Desktop/silver4490/Desktop/Long V1 Envelope O-Glycosylation Story/V1 Insert Los Alamos Blast/Compendium 2016/netOGlyc-4.0.0.13/Hiren_SequencesNotRun_116/"
input = open(inputAlignmentPath + inputAlignment, "r")
lines = input.readlines() # The readline() function returns a list of lines in the file.
input.close()

fullOGlycDict = {}
lines = lines[5:]
for line in lines:
	stripped = line.strip('\n')
	positiveCuratedLine = stripped.split('\t')
	modifiedKey = positiveCuratedLine[0].replace("x", "")
	newlyModifiedKey = modifiedKey.replace("X", "")
	filler = newlyModifiedKey.replace(".", "_")
	particularKey = ">"+filler.upper()
	values = [positiveCuratedLine[3], positiveCuratedLine[5]]
	if particularKey not in fullOGlycDict.keys():
		fullOGlycDict[particularKey] = [values]
	else:
		fullOGlycDict[particularKey].append(values)

i=0
for value in fullOGlycDict.keys():
	# print(value)
	if value in oldNameKey.values():
		i+=1

exclusiveOGlycDict = {}
for key in newlyAlignedSequences.keys():
	exclusiveOGlycDict[key] = fullOGlycDict[key]

# Document predicted O-Glycosylation sites within the V1 Domain
file = open('V1_O-glycosylation.txt', 'w')
mostOutstandingGlycV1Dict = {}
underCutoffDict = {}
outsideV1Dict = {}
noV1Glycosylation = {}
intentionallyRemoved = []
i = 0
for seq in exclusiveOGlycDict:
	cutoff = 0.5
	lowerBound = v1DomainIndices[seq][0]
	upperBound = v1DomainIndices[seq][1]
	for glycSite in exclusiveOGlycDict[seq]:
		infoForLater = [glycSite[0], glycSite[1]]
		itemInQuestion = glycSite[0]
		if int(lowerBound) <= int(itemInQuestion) <= int(upperBound):
			if float(glycSite[1]) >= float(cutoff):
				if seq not in mostOutstandingGlycV1Dict.keys():
					mostOutstandingGlycV1Dict[seq] = [v1Length[seq], infoForLater]
				else:
					mostOutstandingGlycV1Dict[seq].append(infoForLater)
			else:
				if seq not in underCutoffDict.keys():
					underCutoffDict[seq] = [infoForLater]
				elif seq in underCutoffDict.keys():
					underCutoffDict[seq].append(infoForLater)
		else:
			if seq not in outsideV1Dict:
				outsideV1Dict[seq] = [infoForLater]
			elif seq in outsideV1Dict:
				outsideV1Dict[seq].append(infoForLater)
# print(mostOutstandingGlycV1Dict.keys())
# j=0 ########THIS WAS A TEST to see if there was actually a problem. There's not.
# i=0
# for item in mostOutstandingGlycV1Dict.keys():
# 	if item in oldNameKey.keys():
# 		i += 1
# 	else:
# 		j += 1
# print("i=", i, "j=", j)
# print(len(mostOutstandingGlycV1Dict)) #####NOTE: Change the cutoff above to 0.5 and then you will have all of the data desired under "mostOutstandingGlycV1Dict". It will be presented as such: {'>AC_RW_06_175016_JN977603': [27, ['131', '0.188035'], ['136', '0.381907'], ['141', '0.329033'], ['143', '0.667676'], ['147', '0.157004']]}
# print(len(oldNameKey))

##### Use oldNameKey to recover the original names of the sequences as published in Los Alamos
i=0
aminoAcidScoreList = {}
updatedNamesV1Length = {}
updatedNamesSerineThreonineCount = {}
for key in mostOutstandingGlycV1Dict.keys():
	# print(oldNameKey[value] + '\n' + mostOutstandingGlycV1Dict[value])
	newKeyFromLOSALAMOS = oldNameKey[key]
	aminoAcidScoreList[newKeyFromLOSALAMOS] = mostOutstandingGlycV1Dict[key][1:]
	updatedNamesV1Length[newKeyFromLOSALAMOS] = v1Length[key]
	updatedNamesSerineThreonineCount[newKeyFromLOSALAMOS] = serineThreonineCount[key]


for seq in mostOutstandingGlycV1Dict:
	if seq in underCutoffDict:
		del[underCutoffDict[seq]]
	if seq in outsideV1Dict:
		del[outsideV1Dict[seq]]

leastOutstandingGlycV1Dict = {}
for seq in underCutoffDict:
	if seq not in leastOutstandingGlycV1Dict:
		leastOutstandingGlycV1Dict[seq] = [underCutoffDict[seq]]
for seq in outsideV1Dict:
	if seq not in leastOutstandingGlycV1Dict:
		leastOutstandingGlycV1Dict[seq] = outsideV1Dict[seq]

# for item in leastOutstandingGlycV1Dict:
# 	print(leastOutstandingGlycV1Dict[item], '\n')

##########################THIS IS WHERE THE NEW INFORMATION ENDS
##########################

print("There are", len(mostOutstandingGlycV1Dict.keys()), "sequences with at least one O-Glycosylation site scored above", cutoff, ".")
print("There are", len(intentionallyRemoved), "sequences that were analyzed for O-Glycosylation but excluded because those sequences contained at least one unusual character.")
# print("There are", len(overlap), "sequences that did not have O-glycosylation that were in the initial exclusion. The sum of the overlap and unusual character sequences", "(",int(len(overlap)+len(intentionallyRemoved)),")" , "should be equal to", len(removedSequences))

pertinentData = []
cutoffTwo = 0.5
for sequence in mostOutstandingGlycV1Dict:
	i = 1
	generalOGlycScoreList = []
	allOGlycScore = []
	while i < len(mostOutstandingGlycV1Dict[sequence]):
		score = float(mostOutstandingGlycV1Dict[sequence][i][1])
		allOGlycScore.append(score)
		if score >= cutoffTwo:
			generalOGlycScoreList.append(score)
		i += 1
	nameReversal = oldNameKey[sequence]
	mean = numpy.mean(allOGlycScore)
	std = numpy.std(allOGlycScore)
	v1SequenceGlycosylation = [nameReversal, updatedNamesV1Length[nameReversal], updatedNamesSerineThreonineCount[nameReversal], len(generalOGlycScoreList), str(round(mean, 4)) + " (" + str(round(std, 4)) + ")", aminoAcidScoreList[nameReversal][0][0], aminoAcidScoreList[nameReversal][0][1]]
	pertinentData.append(v1SequenceGlycosylation)


allMoreThan9 = []   #######THIS IS FOR GATING ON VERY HIGH VALUED SEQUENCES
for aminoAcid in aminoAcidScoreList:
	print(aminoAcid)
	scores = aminoAcidScoreList[aminoAcid]
	j=0
	i = 0	
	newScores = []
	for score in scores:
		if float(score[1]) >= float(0.8):
			i += 1
		newScores.append(float(score[1]))
		j+=1
	print("i =", i)
	print("j =", j)	
	print(len(scores))
	nameReversal = aminoAcid
	mean = numpy.mean(newScores)
	std = numpy.std(newScores)
	if i == j:
		forAppending = [nameReversal, updatedNamesV1Length[nameReversal], updatedNamesSerineThreonineCount[nameReversal], len(scores), str(round(mean, 4)) + " (" + str(round(std, 4)) + ")", aminoAcidScoreList[nameReversal][0][0], aminoAcidScoreList[nameReversal][0][1]]
		allMoreThan9.append(forAppending)
	else:
		pass

pirate = Workbook()
sheet = pirate.active
sheet.title = "Sequences with a value >0.5"

# thick_border = Border(left=Side(style='thin'), 
#                      right=Side(style='thin'), 
#                      top=Side(style='thin'), 
#                      bottom=Side(style='thick'))


cell = sheet.cell(row=1, column=1)
cell.value = "Sequence Name"
cell = sheet.cell(row=1, column=2)
cell.value = "V1 Length"
cell = sheet.cell(row=1, column=3)
cell.value = "# of S/T in V1"
cell = sheet.cell(row=1, column=4)
cell.value = "# of scored S/T in V1 (>0.5)"
cell = sheet.cell(row=1, column=5)
cell.value = "Mean score (Standard Deviation)"
cell = sheet.cell(row=1, column=6)
cell.value = "Amino Acid #"
cell = sheet.cell(row=1, column=7)
cell.value = "Score"

r=2
i = 0
j=2
for thing in allMoreThan9:
	title = allMoreThan9[i][0]
	v1length = allMoreThan9[i][1]
	serinesThreonines = allMoreThan9[i][2]
	oglycpredictNumber = allMoreThan9[i][3]
	oglycpredictMeanSTD = allMoreThan9[i][4]
	cell = sheet.cell(row=r, column=1)
	cell.value = title
	cell = sheet.cell(row=r, column=2)
	cell.value = v1length	
	cell = sheet.cell(row=r, column=3)
	cell.value = serinesThreonines
	cell = sheet.cell(row=r, column=4)
	cell.value = oglycpredictNumber
	cell = sheet.cell(row=r, column=5)
	cell.value = oglycpredictMeanSTD
	cell = sheet.cell(row=r, column=6)
	cell.value = aminoAcidScoreList[thing[0]][0][0]
	cell = sheet.cell(row=r, column=7)
	cell.value = aminoAcidScoreList[thing[0]][0][1]
	j = 2
	while j <= len(aminoAcidScoreList[thing[0]]):
		r += 1
		cell = sheet.cell(row=r, column=1)
		cell.value = title
		cell = sheet.cell(row=r, column=2)
		cell.value = v1length	
		cell = sheet.cell(row=r, column=3)
		cell.value = serinesThreonines
		cell = sheet.cell(row=r, column=4)
		cell.value = oglycpredictNumber
		cell = sheet.cell(row=r, column=5)
		cell.value = oglycpredictMeanSTD
		cell = sheet.cell(row=r, column=6)
		cell.value = aminoAcidScoreList[thing[0]][j-1][0]
		cell = sheet.cell(row=r, column=7)
		cell.value = aminoAcidScoreList[thing[0]][j-1][1]
		j += 1
	# mergeIdentifier = 'A' + str(r-len(aminoAcidScoreList[thing[0]])+1) + ":" + "A" + str(r)
	# mergeIdentifier2 = 'B' + str(r-len(aminoAcidScoreList[thing[0]])+1) + ":" + "B" + str(r)
	# mergeIdentifier3 = 'C' + str(r-len(aminoAcidScoreList[thing[0]])+1) + ":" + "C" + str(r)
	# mergeIdentifier4 = 'D' + str(r-len(aminoAcidScoreList[thing[0]])+1) + ":" + "D" + str(r)
	# mergeIdentifier5 = 'E' + str(r-len(aminoAcidScoreList[thing[0]])+1) + ":" + "E" + str(r)
	# print(mergeIdentifier)
	# sheet.merge_cells(mergeIdentifier)
	# sheet.merge_cells(mergeIdentifier2)
	# sheet.merge_cells(mergeIdentifier3)
	# sheet.merge_cells(mergeIdentifier4)
	# sheet.merge_cells(mergeIdentifier5)
	r += 1
	i += 1

newInputPath = '/Users/newumuser/Desktop/silver4490/Desktop/'
pirate.save(newInputPath + "20180802_ONLYHIGHVALUES" + ".xlsx")

raise SystemExit

for eachList in pertinentData:
	outstring = ''
	for eachItem in eachList:
		outstring += str(eachItem) + "," 
	output.write(outstring)
	i = 2
	newstring = ''
	while i <= len(aminoAcidScoreList[eachList[0]]):
		newstring = '\n' + eachList[0] + ',' + '' + ',' +'' + ','+ '' + ','+ '' + ',' + str(aminoAcidScoreList[eachList[0]][i-1][0]) + "," + str(aminoAcidScoreList[eachList[0]][i-1][1]) + ','
		output.write(newstring)
		i += 1
	output.write('\n')
	i = 0
output.close

r=2
i = 0
for thing in pertinentData:
	title = pertinentData[i][0]
	v1length = pertinentData[i][1]
	glyc = pertinentData[i][2]
	cell = sheet.cell(row=r, column=1)
	cell.value = title
	cell = sheet.cell(row=r, column=2)
	cell.value = v1length	
	cell = sheet.cell(row=r, column=3)
	cell.value = glyc
	cell = sheet.cell(row=r, column=4)
	cell.value = float((glyc/v1length)*100)
	r += 1
	i += 1


raise SystemExit

# threeDGraph = open('RawData_XYZ_V1Length.txt', 'w')
# for item in pertinentData:
# 	stringItem = str(item)
# 	threeDGraph.write(stringItem + '\n')
# 	threeDGraph.write(stringItem + '\n')

# x_value_V1Length = []
# y_value_beforeTreatment = []
# z_value_NumPosSites = []
# for data in pertinentData:
# 	x_value_V1Length.append(data[1])
# 	y_value_beforeTreatment.append(data[3])
# 	z_value_NumPosSites.append(data[2])

# a=numpy.array(y_value_beforeTreatment)
# s=numpy.isnan(a)
# a[s]=0.0
# y_value_AvgV1Score = []
# for each in a:
# 	y_value_AvgV1Score.append(each)

# from matplotlib import pyplot
# from mpl_toolkits.mplot3d import Axes3D
# import random

# fig = pyplot.figure()
# ax = Axes3D(fig)

# # sequence_containing_x_vals = x_value_V1Length.range(0, 70)
# # sequence_containing_y_vals = y_value_AvgV1Score(range(0, 1.0))
# # sequence_containing_z_vals = z_value_NumPosSites(range(0, 40))

# random.shuffle(x_value_V1Length)
# random.shuffle(y_value_AvgV1Score)
# random.shuffle(z_value_NumPosSites)

# ax.scatter(y_value_AvgV1Score, z_value_NumPosSites, x_value_V1Length)
# ax.set_zlabel('Number of Positive Predictions')
# ax.set_xlabel('V1 Domain Length')
# ax.set_ylabel('V1 Domain Average Score')
# pyplot.show()
cell = sheet.cell(row=1, column=1)
cell.value = "Name"
cell = sheet.cell(row=1, column=2)
cell.value = "V1 Length"
cell = sheet.cell(row=1, column=3)
cell.value = "# of predicted O-glycosylation sites (>0.9)"
cell = sheet.cell(row=1, column=4)
cell.value = "Percent of V1 predicted to be O-glycosylated (>0.9)"

cell2 = ws2.cell(row=1, column=1)
cell2.value = "Length of V1"
cell2 = ws2.cell(row=1, column=2)
cell2.value = "Number of sequences"